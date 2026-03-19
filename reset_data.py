"""
reset_data.py — Limpia los datos de prueba del sistema.

Modos:
  python reset_data.py          -> Limpia solo los REGISTROS (REG/) — conserva catálogos
  python reset_data.py --todo   -> Limpia registros Y regenera catálogos con datos de prueba
  python reset_data.py --cat    -> Limpia solo los catálogos CAT/
"""
import os
import sys
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CAT_DIR = os.path.join(BASE_DIR, "CAT")
REG_DIR = os.path.join(BASE_DIR, "REG")

# Archivos REG y sus hojas de datos
REG_FILES = {
    "RegistroMovMaterial.xlsx":    ["MOV_ALMACEN"],
    "RegistroMovHerramienta.xlsx": ["MOV_HERRA"],
    "Lotes.xlsx":                  ["LOTES"],
    "RegAvance.xlsx":              ["REG_AVANCE"],
}

# Archivos CAT y sus hojas de datos (todos excepto sheets de apoyo)
CAT_FILES = {
    "Materiales.xlsx":    ["MATERIALES"],
    "Herramientas.xlsx":  ["HERRAMIENTAS"],
    "Empleados.xlsx":     ["EMPLEADOS"],
    "Ubicaciones.xlsx":   ["UBICACIONES"],
    "Proveedores.xlsx":   ["PROVEEDORES"],
    "Proyectos.xlsx":     ["PROYECTOS", "MUEBLES", "ORDENES_PRODUCCION"],
    # ETAPAS se conserva siempre (datos de configuracion, no de prueba)
}


def limpiar_hoja(filepath, sheet_name):
    """Elimina todas las filas de datos (conserva la fila de encabezados)."""
    wb = load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        print(f"  [AVISO] Hoja '{sheet_name}' no encontrada en {os.path.basename(filepath)}")
        return
    ws = wb[sheet_name]
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)
        print(f"  Limpiado: {os.path.basename(filepath)} -> {sheet_name} ({max_row-1} filas eliminadas)")
    else:
        print(f"  Sin datos: {os.path.basename(filepath)} -> {sheet_name}")
    wb.save(filepath)


def limpiar_registros():
    print("\n[1] Limpiando archivos de REGISTROS (REG/)...")
    for filename, sheets in REG_FILES.items():
        filepath = os.path.join(REG_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  [AVISO] No encontrado: {filename}")
            continue
        for sheet in sheets:
            limpiar_hoja(filepath, sheet)
    print("  Registros vaciados.")


def limpiar_catalogos():
    print("\n[2] Limpiando archivos de CATALOGOS (CAT/)...")
    for filename, sheets in CAT_FILES.items():
        filepath = os.path.join(CAT_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  [AVISO] No encontrado: {filename}")
            continue
        for sheet in sheets:
            limpiar_hoja(filepath, sheet)
    print("  Catalogos vaciados.")


def confirmar(mensaje):
    resp = input(f"\n{mensaje} (s/N): ").strip().lower()
    return resp in ("s", "si", "y", "yes")


if __name__ == "__main__":
    args = sys.argv[1:]
    print("=" * 50)
    print("  RESET DE DATOS — Sistema Taller Carpinteria")
    print("=" * 50)

    if "--todo" in args:
        if confirmar("Esto eliminara TODOS los datos de prueba (registros + catalogos). Continuar?"):
            limpiar_registros()
            limpiar_catalogos()
            print("\nTodo limpio. Ejecuta setup_excel.py para volver a generar datos de prueba.")
        else:
            print("Cancelado.")

    elif "--cat" in args:
        if confirmar("Esto eliminara los datos de prueba de los CATALOGOS. Continuar?"):
            limpiar_catalogos()
        else:
            print("Cancelado.")

    else:
        print("\nModo por defecto: limpia solo los REGISTROS (REG/).")
        print("Los catalogos (empleados, materiales, etc.) se conservan.\n")
        if confirmar("Eliminar todos los registros de movimientos, prestamos, lotes y avance?"):
            limpiar_registros()
            print("\nRegistros vaciados. Los catalogos permanecen intactos.")
        else:
            print("Cancelado.")

    print()
