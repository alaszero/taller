"""
importar_excel_a_sqlite.py
Importa todos los datos de los archivos Excel (CAT/ y REG/) al SQLite local (taller.db).
- Asigna taller_id = 'rober_lang' a todos los registros
- Borra primero los 2 registros de prueba de MATERIALES (MAT_TEST1, MAT_TEST2)
- Usa INSERT OR IGNORE en tablas con PK para evitar duplicados
- Para REG (movimientos/avances): borra primero si ya hay datos, luego inserta

Uso:  python importar_excel_a_sqlite.py
"""

import sqlite3
import openpyxl
import unicodedata
import os

DB_PATH = "taller.db"
TALLER_ID = "rober_lang"

# Mapa archivo Excel → tabla SQLite → método de inserción
IMPORTACIONES = [
    ("CAT/Materiales.xlsx",             "MATERIALES",    "ignore"),
    ("CAT/Herramientas.xlsx",           "HERRAMIENTAS",  "ignore"),
    ("CAT/Empleados.xlsx",              "EMPLEADOS",     "ignore"),
    ("CAT/Proyectos.xlsx",              "PROYECTOS",     "ignore"),
    ("CAT/Ubicaciones.xlsx",            "UBICACIONES",   "ignore"),
    ("CAT/Proveedores.xlsx",            "PROVEEDORES",   "ignore"),
    ("REG/Lotes.xlsx",                  "LOTES",         "replace"),
    ("REG/RegAvance.xlsx",              "REG_AVANCE",    "replace"),
    ("REG/RegistroMovHerramienta.xlsx", "MOV_HERRA",     "replace"),
    ("REG/RegistroMovMaterial.xlsx",    "MOV_ALMACEN",   "replace"),
]


def normalize(s: str) -> str:
    """Normaliza string: quita acentos y pasa a minúsculas para comparación."""
    if not s:
        return ""
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii").lower().strip()


def get_sqlite_cols(conn: sqlite3.Connection, table: str) -> list[str]:
    return [r[1] for r in conn.execute(f'PRAGMA table_info("{table}")').fetchall()]


def map_headers(excel_headers: list, sqlite_cols: list) -> dict:
    """
    Devuelve {idx_excel: nombre_col_sqlite} para cada header Excel
    que tenga match en SQLite (comparación normalizada).
    """
    norm_map = {normalize(c): c for c in sqlite_cols}
    mapping = {}
    for i, h in enumerate(excel_headers):
        if h is None:
            continue
        norm_h = normalize(str(h))
        if norm_h in norm_map:
            mapping[i] = norm_map[norm_h]
    return mapping


def import_table(conn: sqlite3.Connection, xlsx_path: str, table: str, mode: str):
    if not os.path.exists(xlsx_path):
        print(f"  [SKIP] No encontrado: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Leer cabeceras
    excel_headers = [cell.value for cell in ws[1]]
    sqlite_cols = get_sqlite_cols(conn, table)
    col_mapping = map_headers(excel_headers, sqlite_cols)

    if not col_mapping:
        print(f"  [ERROR] Sin columnas mapeadas para {table}. Headers Excel: {excel_headers}")
        return

    # En modo replace borramos antes
    if mode == "replace":
        deleted = conn.execute(
            f'DELETE FROM "{table}" WHERE taller_id=? OR taller_id IS NULL OR taller_id=\'\'',
            (TALLER_ID,)
        ).rowcount
        if deleted:
            print(f"  Borrados {deleted} registros previos de {table}")

    # Construir INSERT
    dest_cols = list(col_mapping.values()) + ["taller_id"]
    col_list  = ", ".join(f'"{c}"' for c in dest_cols)
    placeholders = ", ".join("?" for _ in dest_cols)

    if mode == "ignore":
        sql = f'INSERT OR IGNORE INTO "{table}" ({col_list}) VALUES ({placeholders})'
    else:
        sql = f'INSERT INTO "{table}" ({col_list}) VALUES ({placeholders})'

    inserted = 0
    skipped  = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Ignorar filas completamente vacías
        if all(v is None for v in row):
            continue

        values = []
        for i in range(len(excel_headers)):
            if i in col_mapping:
                val = row[i] if i < len(row) else None
                # Convertir fechas/números a str
                if val is None:
                    values.append("")
                else:
                    values.append(str(val).strip())
        values.append(TALLER_ID)

        try:
            cur = conn.execute(sql, values)
            if cur.rowcount > 0:
                inserted += 1
            else:
                skipped += 1
        except sqlite3.IntegrityError:
            skipped += 1
        except Exception as e:
            print(f"  [WARN] Fila ignorada ({e}): {values}")

    conn.commit()
    print(f"  {table}: {inserted} insertados, {skipped} omitidos")


def main():
    print(f"Conectando a {DB_PATH}...")
    conn = sqlite3.connect(DB_PATH)

    # Borrar registros de prueba
    test_deleted = conn.execute(
        "DELETE FROM MATERIALES WHERE Codigo_mat IN ('MAT_TEST1', 'MAT_TEST2')"
    ).rowcount
    if test_deleted:
        print(f"Borrados {test_deleted} materiales de prueba (MAT_TEST1, MAT_TEST2)\n")
    conn.commit()

    # Importar cada archivo
    for xlsx_path, table, mode in IMPORTACIONES:
        print(f"Importando {xlsx_path} -> {table} (modo: {mode})")
        import_table(conn, xlsx_path, table, mode)

    # Verificación final
    print("\n=== VERIFICACIÓN FINAL ===")
    for _, table, _ in IMPORTACIONES:
        count = conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
        print(f"  {table}: {count} registros totales")

    conn.close()
    print("\nImportación completada.")


if __name__ == "__main__":
    main()
