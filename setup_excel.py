"""
setup_excel.py — Genera todos los archivos Excel del sistema de taller de carpintería.
Los datos de prueba se marcan con fondo amarillo para identificarlos fácilmente.
Ejecutar: python setup_excel.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CAT_DIR = os.path.join(BASE_DIR, "CAT")
REG_DIR = os.path.join(BASE_DIR, "REG")

# ── Estilos ──────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT  = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(horizontal="left", vertical="center")
PRUEBA_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")  # Amarillo = dato de prueba
ALT_FILL   = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")   # Azul claro alterno

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

def style_header(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = HDR_ALIGN
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 28

def style_row(ws, row_num, n_cols, is_sample=True):
    fill = PRUEBA_FILL if is_sample else ALT_FILL
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.font = DATA_FONT
        c.fill = fill
        c.alignment = DATA_ALIGN
        c.border = THIN_BORDER
    ws.row_dimensions[row_num].height = 18

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def add_note(ws, note_text):
    """Añade una nota en la celda A1 de una hoja auxiliar."""
    pass

# ── CAT/Materiales.xlsx ───────────────────────────────────────────────────────
def create_materiales():
    wb = Workbook()
    ws = wb.active
    ws.title = "MATERIALES"
    ws.sheet_view.showGridLines = False

    headers = ["Codigo_mat", "Tipo", "Descripción", "Unidad", "Stock_min"]
    style_header(ws, headers)

    data = [
        ["MAT001", "Tablero",    "MDF 15mm 1.22x2.44m",           "pza",  10],
        ["MAT002", "Tablero",    "MDF 9mm 1.22x2.44m",            "pza",   5],
        ["MAT003", "Tablero",    "Triplay 9mm 1.22x2.44m",         "pza",   5],
        ["MAT004", "Tablero",    "Melamínico Blanco 15mm 1.22x2.44m","pza", 8],
        ["MAT005", "Herraje",    "Bisagra 35mm Blum",              "pza", 100],
        ["MAT006", "Herraje",    "Corredera 45cm Grass",           "par",  50],
        ["MAT007", "Herraje",    "Jalador barra acero 128mm",      "pza",  30],
        ["MAT008", "Acabado",    "Laca NC transparente mate",      "lt",   10],
        ["MAT009", "Acabado",    "Sellador nitrocelulosa",         "lt",    8],
        ["MAT010", "Acabado",    "Thinner estándar",               "lt",   20],
        ["MAT011", "Consumible", "Lija 120 pliego",                "pza",  50],
        ["MAT012", "Consumible", "Masilla plástica",               "kg",    5],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [14, 14, 38, 8, 12])
    freeze_filter(ws)

    # Hoja de resumen de stock (se actualiza desde la app)
    ws2 = wb.create_sheet("STOCK_RESUMEN")
    ws2.sheet_view.showGridLines = False
    stock_hdrs = ["Codigo_mat", "Descripción", "Unidad", "Stock_min", "Stock_actual", "Alerta"]
    style_header(ws2, stock_hdrs)
    ws2.append(["(Actualizar desde la app Flask o ejecutar recálculo)", "", "", "", "", ""])
    set_col_widths(ws2, [14, 38, 8, 12, 14, 10])

    wb.save(os.path.join(CAT_DIR, "Materiales.xlsx"))
    print("  OK CAT/Materiales.xlsx")


# ── CAT/Herramientas.xlsx ─────────────────────────────────────────────────────
def create_herramientas():
    wb = Workbook()
    ws = wb.active
    ws.title = "HERRAMIENTAS"
    ws.sheet_view.showGridLines = False

    headers = ["ID_Herramienta","Herramienta","Marca","Modelo","Tipo",
               "NS","Categoria","Ubicación_Base","Estado","Responsable_Actual",
               "Fecha_Alta","Observaciones"]
    style_header(ws, headers)

    data = [
        ["HER001","Sierra circular",   "Makita",    "5007MG",    "Eléctrica",  "SN001","Corte",    "Taller Principal","Disponible","",          "2022-01-15",""],
        ["HER002","Router/Fresadora",  "Bosch",     "POF1400",   "Eléctrica",  "SN002","Maquinado","Taller Principal","Disponible","",          "2022-01-15",""],
        ["HER003","Lijadora orbital",  "Makita",    "BO5041",    "Eléctrica",  "SN003","Acabado",  "Bodega",          "Disponible","",          "2022-03-10",""],
        ["HER004","Taladro",           "DeWalt",    "DWD112",    "Eléctrica",  "SN004","Ensamble", "Taller Principal","Prestada",  "EMP002",    "2021-06-20",""],
        ["HER005","Pistola laqueadora","Sata",      "Jet 3000",  "Neumática",  "SN005","Acabado",  "Zona Lacado",     "Disponible","",          "2022-06-05",""],
        ["HER006","Compresor 100L",    "Campbell",  "VT6271",    "Neumática",  "SN006","General",  "Bodega",          "Disponible","",          "2021-01-10",""],
        ["HER007","Cepillo eléctrico", "Makita",    "KP0810",    "Eléctrica",  "SN007","Maquinado","Taller Principal","Disponible","",          "2022-08-15",""],
        ["HER008","Caladora",          "Bosch",     "PST900",    "Eléctrica",  "SN008","Corte",    "Taller Principal","Disponible","",          "2022-09-01",""],
        ["HER009","Escuadra digital",  "Starrett",  "-",         "Medición",   "SN009","General",  "Taller Principal","Disponible","",          "2023-01-15",""],
        ["HER010","Nivel láser",       "Bosch",     "GLL3-80",   "Medición",   "SN010","General",  "Taller Principal","Disponible","",          "2023-02-20",""],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [14,20,12,12,12,10,12,18,12,14,12,20])
    freeze_filter(ws)
    wb.save(os.path.join(CAT_DIR, "Herramientas.xlsx"))
    print("  OK CAT/Herramientas.xlsx")


# ── CAT/Empleados.xlsx ────────────────────────────────────────────────────────
def create_empleados():
    wb = Workbook()
    ws = wb.active
    ws.title = "EMPLEADOS"
    ws.sheet_view.showGridLines = False

    headers = ["EmpleadoID","Nombre","Apellido","Alias","Area","Puesto",
               "Especialidad","NivelExperiencia","FechaIngreso","SalarioHora",
               "CostoHoraReal","Activo","SupervisorID","Observaciones"]
    style_header(ws, headers)

    data = [
        ["EMP001","Carlos",   "Ramírez", "El Flaco", "Carpintería","Carpintero","Armado",      4,"2019-03-15",120,145,"Sí","EMP007",""],
        ["EMP002","Miguel",   "Torres",  "Miguelito","Carpintería","Carpintero","Corte",        3,"2021-01-10",100,120,"Sí","EMP007",""],
        ["EMP003","Roberto",  "Sánchez", "Beto",     "Carpintería","Carpintero","Maquinado",    5,"2017-06-20",140,168,"Sí","EMP007",""],
        ["EMP004","Luis",     "Hernández","Güero",   "Lacado",     "Laqueador", "Laqueado",     4,"2020-04-15",115,138,"Sí","EMP008",""],
        ["EMP005","Pedro",    "Martínez","Pepe",     "Lacado",     "Laqueador", "Masillado",    2,"2023-02-01", 90,108,"Sí","EMP008",""],
        ["EMP006","Antonio",  "López",   "Toño",     "Carpintería","Carpintero","Instalación",  3,"2022-07-15",100,120,"Sí","EMP007",""],
        ["EMP007","Javier",   "Morales", "El Jefe",  "Carpintería","Supervisor","General",      5,"2015-01-05",180,220,"Sí","",     "Supervisor Carpintería"],
        ["EMP008","Ana",      "García",  "Ana",      "Lacado",     "Supervisor","General",      4,"2018-09-10",160,195,"Sí","",     "Supervisora Lacado"],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [10,12,14,12,14,14,14,16,14,14,14,8,12,22])
    freeze_filter(ws)
    wb.save(os.path.join(CAT_DIR, "Empleados.xlsx"))
    print("  OK CAT/Empleados.xlsx")


# ── CAT/Ubicaciones.xlsx ──────────────────────────────────────────────────────
def create_ubicaciones():
    wb = Workbook()
    ws = wb.active
    ws.title = "UBICACIONES"
    ws.sheet_view.showGridLines = False

    headers = ["ID_Ubic","Zona","Descripción","Tipo"]
    style_header(ws, headers)

    data = [
        ["UB001","Almacén","Bodega principal",         "Almacenamiento"],
        ["UB002","Taller", "Área de corte",            "Producción"],
        ["UB003","Taller", "Área de armado",           "Producción"],
        ["UB004","Lacado", "Zona de laqueado",         "Acabado"],
        ["UB005","Exterior","Patio trasero",           "Almacenamiento"],
        ["UB006","Oficina","Área administrativa",      "Administrativo"],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [10,14,28,18])
    freeze_filter(ws)
    wb.save(os.path.join(CAT_DIR, "Ubicaciones.xlsx"))
    print("  OK CAT/Ubicaciones.xlsx")


# ── CAT/Proveedores.xlsx ──────────────────────────────────────────────────────
def create_proveedores():
    wb = Workbook()
    ws = wb.active
    ws.title = "PROVEEDORES"
    ws.sheet_view.showGridLines = False

    headers = ["ProveedorID","Nombre","Telefono","Email","RFC","Contacto","Categoria","Observaciones"]
    style_header(ws, headers)

    data = [
        ["PROV001","Maderas del Norte",     "5555-1234","maderas@norte.com",   "MDN010101ABC","Ing. Castro",  "Madera y tableros",""],
        ["PROV002","Herrajes Industriales", "5555-2345","herrajes@ind.com",    "HIN020202DEF","Lic. Pérez",   "Herrajes y accesorios",""],
        ["PROV003","Lacas y Acabados SA",   "5555-3456","lacas@acabados.com",  "LAS030303GHI","Sr. Guzmán",   "Acabados y químicos",""],
        ["PROV004","Materiales Express",    "5555-4567","mat@express.com",     "MEX040404JKL","Sr. Domínguez","General","Entrega rápida"],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [12,24,14,26,18,16,22,22])
    freeze_filter(ws)
    wb.save(os.path.join(CAT_DIR, "Proveedores.xlsx"))
    print("  OK CAT/Proveedores.xlsx")


# ── CAT/Proyectos.xlsx (4 hojas) ──────────────────────────────────────────────
def create_proyectos():
    wb = Workbook()

    # ── Hoja 1: PROYECTOS ──
    ws1 = wb.active
    ws1.title = "PROYECTOS"
    ws1.sheet_view.showGridLines = False
    h1 = ["ID_Proyecto","Cliente","Nombre_Obra","Inicio","Fin","Encargado","Estado"]
    style_header(ws1, h1)
    proy_data = [
        ["PROY001","Familia Rodríguez","Cocina integral + closets",  "2026-01-15","2026-04-30","EMP007","En proceso"],
        ["PROY002","Oficinas ABC",     "Muebles de oficina modulares","2026-02-01","2026-05-15","EMP007","En proceso"],
    ]
    for i, row in enumerate(proy_data, 2):
        for j, val in enumerate(row, 1):
            ws1.cell(row=i, column=j, value=val)
        style_row(ws1, i, len(h1))
    set_col_widths(ws1, [12,22,30,12,12,12,14])
    freeze_filter(ws1)

    # ── Hoja 2: MUEBLES ──
    ws2 = wb.create_sheet("MUEBLES")
    ws2.sheet_view.showGridLines = False
    h2 = ["ID_Mueble","ID_Proyecto","Mueble","Cantidad","Tipo","Area","Observación"]
    style_header(ws2, h2)
    muebles_data = [
        ["MUE001","PROY001","Cocina integral",          1,"Cocina",    "Carpintería",""],
        ["MUE002","PROY001","Closet habitación principal",2,"Closet",  "Carpintería","Puertas corredizas"],
        ["MUE003","PROY001","Closet habitación 2",      1,"Closet",    "Carpintería",""],
        ["MUE004","PROY002","Escritorio gerencial",      3,"Escritorio","Carpintería","Con cajonera"],
        ["MUE005","PROY002","Mueble archivero",          5,"Archivero", "Carpintería",""],
        ["MUE006","PROY002","Mesa de reuniones",         1,"Mesa",      "Carpintería","Para 10 personas"],
    ]
    for i, row in enumerate(muebles_data, 2):
        for j, val in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=val)
        style_row(ws2, i, len(h2))
    set_col_widths(ws2, [10,12,28,10,14,14,24])
    freeze_filter(ws2)

    # ── Hoja 3: ORDENES_PRODUCCION ──
    ws3 = wb.create_sheet("ORDENES_PRODUCCION")
    ws3.sheet_view.showGridLines = False
    h3 = ["ID_OP","ID_Mueble","Etapa","Cantidad",
          "FechaInicioPlaneada","FechaFinalPlaneada",
          "FechaInicioReal","FechaFinalReal","Responsable"]
    style_header(ws3, h3)
    op_data = [
        ["OP001","MUE001","Corte",   1,"2026-01-20","2026-01-25","2026-01-21","2026-01-24","EMP002"],
        ["OP002","MUE001","Armado",  1,"2026-01-26","2026-02-05","2026-01-26","",          "EMP001"],
        ["OP003","MUE002","Corte",   2,"2026-01-22","2026-01-27","2026-01-22","2026-01-26","EMP002"],
        ["OP004","MUE002","Armado",  2,"2026-01-28","2026-02-08","2026-01-29","",          "EMP001"],
        ["OP005","MUE003","Corte",   1,"2026-02-10","2026-02-14","",          "",          "EMP002"],
        ["OP006","MUE004","Corte",   3,"2026-02-05","2026-02-10","2026-02-05","2026-02-09","EMP003"],
    ]
    for i, row in enumerate(op_data, 2):
        for j, val in enumerate(row, 1):
            ws3.cell(row=i, column=j, value=val)
        style_row(ws3, i, len(h3))
    set_col_widths(ws3, [8,10,14,10,20,20,18,18,12])
    freeze_filter(ws3)

    # ── Hoja 4: ETAPAS ──
    ws4 = wb.create_sheet("ETAPAS")
    ws4.sheet_view.showGridLines = False
    h4 = ["Etapa","Orden","Area"]
    style_header(ws4, h4)
    etapas_data = [
        ["Corte",         1,"Carpintería"],
        ["Maquinado",     2,"Carpintería"],
        ["Armado",        3,"Carpintería"],
        ["Masillado",     4,"Lacado"],
        ["Lijado",        5,"Lacado"],
        ["Sellado",       6,"Lacado"],
        ["Laqueado base", 7,"Lacado"],
        ["Lijado fino",   8,"Lacado"],
        ["Laqueado final",9,"Lacado"],
        ["Instalación",  10,"Carpintería"],
    ]
    for i, row in enumerate(etapas_data, 2):
        for j, val in enumerate(row, 1):
            ws4.cell(row=i, column=j, value=val)
        style_row(ws4, i, len(h4))
    set_col_widths(ws4, [18,8,14])
    freeze_filter(ws4)

    wb.save(os.path.join(CAT_DIR, "Proyectos.xlsx"))
    print("  OK CAT/Proyectos.xlsx  (4 hojas)")


# ── REG/RegistroMovMaterial.xlsx ──────────────────────────────────────────────
def create_mov_material():
    wb = Workbook()
    ws = wb.active
    ws.title = "MOV_ALMACEN"
    ws.sheet_view.showGridLines = False

    headers = ["Fecha","Codigo_mat","Tipo","Cantidad","Unidad",
               "ID_Proyecto","Mueble","ID_Trab","ID_Ubic_Origen","ID_Ubic_Destino"]
    style_header(ws, headers)

    data = [
        ["2026-01-21","MAT001","Salida",  3,"pza","PROY001","Cocina integral",          "EMP002","UB001","UB002"],
        ["2026-01-21","MAT005","Salida", 20,"pza","PROY001","Cocina integral",          "EMP002","UB001","UB003"],
        ["2026-01-22","MAT009","Salida",  5,"lt", "PROY001","Cocina integral",          "EMP004","UB001","UB004"],
        ["2026-02-01","MAT001","Entrada",20,"pza","",        "",                         "EMP007","PROV001","UB001"],
        ["2026-02-05","MAT005","Salida", 15,"pza","PROY001","Closet habitación principal","EMP001","UB001","UB003"],
        ["2026-02-06","MAT008","Salida",  8,"lt", "PROY001","Closet habitación principal","EMP004","UB001","UB004"],
        ["2026-02-10","MAT002","Salida",  4,"pza","PROY002","Escritorio gerencial",     "EMP003","UB001","UB002"],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [14,12,10,10,8,12,28,10,16,16])
    freeze_filter(ws)
    wb.save(os.path.join(REG_DIR, "RegistroMovMaterial.xlsx"))
    print("  OK REG/RegistroMovMaterial.xlsx")


# ── REG/RegistroMovHerramienta.xlsx ──────────────────────────────────────────
def create_mov_herramienta():
    wb = Workbook()
    ws = wb.active
    ws.title = "MOV_HERRA"
    ws.sheet_view.showGridLines = False

    headers = ["ID_Prestamo","FechaSalida","ID_Herramienta","Herramienta",
               "Responsable","Nombre_Responsable","Proyecto","FechaDevolucion"]
    style_header(ws, headers)

    data = [
        ["PREST001","2026-01-21","HER001","Sierra circular",   "EMP002","Miguel Torres",   "PROY001","2026-01-25"],
        ["PREST002","2026-01-22","HER004","Taladro",           "EMP001","Carlos Ramírez",  "PROY001",""],
        ["PREST003","2026-02-01","HER003","Lijadora orbital",  "EMP004","Luis Hernández",  "PROY001","2026-02-03"],
        ["PREST004","2026-02-05","HER007","Cepillo eléctrico", "EMP003","Roberto Sánchez", "PROY001",""],
        ["PREST005","2026-02-10","HER002","Router/Fresadora",  "EMP003","Roberto Sánchez", "PROY002",""],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [12,14,16,22,12,20,12,16])
    freeze_filter(ws)
    wb.save(os.path.join(REG_DIR, "RegistroMovHerramienta.xlsx"))
    print("  OK REG/RegistroMovHerramienta.xlsx")


# ── REG/Lotes.xlsx ────────────────────────────────────────────────────────────
def create_lotes():
    wb = Workbook()
    ws = wb.active
    ws.title = "LOTES"
    ws.sheet_view.showGridLines = False

    headers = ["LoteID","MaterialID","FechaCompra","FechaCaducidad",
               "CantidadInicial","CantidadDisponible","UbicacionID",
               "CostoUnitario","ProveedorID","EstadoLote"]
    style_header(ws, headers)

    data = [
        ["LOT001","MAT001","2026-01-10","",50,47,"UB001",285.00,"PROV001","Activo"],
        ["LOT002","MAT002","2026-01-10","",30,30,"UB001",220.00,"PROV001","Activo"],
        ["LOT003","MAT005","2026-01-10","",200,180,"UB001",45.50,"PROV002","Activo"],
        ["LOT004","MAT006","2026-01-12","",100,100,"UB001",125.00,"PROV002","Activo"],
        ["LOT005","MAT008","2026-01-15","",30,22,"UB001",180.00,"PROV003","Activo"],
        ["LOT006","MAT009","2026-01-15","",20,15,"UB001",165.00,"PROV003","Activo"],
        ["LOT007","MAT010","2026-01-15","",40,40,"UB001", 52.00,"PROV003","Activo"],
        ["LOT008","MAT011","2026-01-20","",100,100,"UB001", 8.50,"PROV004","Activo"],
        ["LOT009","MAT001","2026-02-01","",20,20,"UB001",290.00,"PROV001","Activo"],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [10,12,14,14,16,18,12,14,12,12])
    freeze_filter(ws)
    wb.save(os.path.join(REG_DIR, "Lotes.xlsx"))
    print("  OK REG/Lotes.xlsx")


# ── REG/RegAvance.xlsx ────────────────────────────────────────────────────────
def create_reg_avance():
    wb = Workbook()
    ws = wb.active
    ws.title = "REG_AVANCE"
    ws.sheet_view.showGridLines = False

    headers = ["Fecha","ID_OP","EmpleadoID","Etapa","Estado","Horas","Piezas"]
    style_header(ws, headers)

    data = [
        ["2026-01-21","OP001","EMP002","Corte",   "Completado",8, 1],
        ["2026-01-22","OP003","EMP002","Corte",   "Completado",8, 2],
        ["2026-01-24","OP003","EMP002","Corte",   "Completado",8, 2],
        ["2026-01-26","OP002","EMP001","Armado",  "En proceso", 6,""],
        ["2026-01-29","OP004","EMP001","Armado",  "En proceso", 8,""],
        ["2026-02-05","OP006","EMP003","Corte",   "Completado",8, 3],
        ["2026-02-06","OP006","EMP003","Corte",   "Completado",6, 3],
    ]
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        style_row(ws, i, len(headers))

    set_col_widths(ws, [14,8,12,16,14,8,8])
    freeze_filter(ws)
    wb.save(os.path.join(REG_DIR, "RegAvance.xlsx"))
    print("  OK REG/RegAvance.xlsx")


# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n=== Generando archivos Excel del Sistema de Taller ===\n")
    print("Catalogos:")
    create_materiales()
    create_herramientas()
    create_empleados()
    create_ubicaciones()
    create_proveedores()
    create_proyectos()
    print("\nRegistros:")
    create_mov_material()
    create_mov_herramienta()
    create_lotes()
    create_reg_avance()
    print("\nTodos los archivos generados en D:/Proyectos/")
    print("  Nota: Las filas en AMARILLO son datos de prueba.")
    print("  Para borrarlos, ejecuta: python reset_data.py\n")
