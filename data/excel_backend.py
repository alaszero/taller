"""
data/excel_backend.py — Capa de acceso a datos Excel/local.
Expone la misma interfaz pública que sqlite_backend.py para que
los blueprints funcionen sin cambios con ambos backends.
"""
import os
import json
import sqlite3
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from flask import abort, session
from werkzeug.security import generate_password_hash

import db as _db
from core.config import DEFAULT_SECCIONES, _CROSS_TALLER_TABLES
from core.helpers import safe_float

# ── Directorios ──────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAT      = os.path.join(BASE_DIR, "CAT")
REG      = os.path.join(BASE_DIR, "REG")
DB_PATH  = os.path.join(BASE_DIR, "users.db")   # solo para USUARIOS

# Mapeo: tabla -> (filepath, sheet_name)
_TABLE_MAP = {
    "MATERIALES":         (lambda: os.path.join(CAT, "Materiales.xlsx"),            "MATERIALES"),
    "HERRAMIENTAS":       (lambda: os.path.join(CAT, "Herramientas.xlsx"),           "HERRAMIENTAS"),
    "EMPLEADOS":          (lambda: os.path.join(CAT, "Empleados.xlsx"),              "EMPLEADOS"),
    "UBICACIONES":        (lambda: os.path.join(CAT, "Ubicaciones.xlsx"),            "UBICACIONES"),
    "PROVEEDORES":        (lambda: os.path.join(CAT, "Proveedores.xlsx"),            "PROVEEDORES"),
    "PROYECTOS":          (lambda: os.path.join(CAT, "Proyectos.xlsx"),              "PROYECTOS"),
    "MUEBLES":            (lambda: os.path.join(CAT, "Proyectos.xlsx"),              "MUEBLES"),
    "ORDENES_PRODUCCION": (lambda: os.path.join(CAT, "Proyectos.xlsx"),              "ORDENES_PRODUCCION"),
    "ETAPAS":             (lambda: os.path.join(CAT, "Proyectos.xlsx"),              "ETAPAS"),
    "MOV_ALMACEN":        (lambda: os.path.join(REG, "RegistroMovMaterial.xlsx"),    "MOV_ALMACEN"),
    "MOV_HERRA":          (lambda: os.path.join(REG, "RegistroMovHerramienta.xlsx"), "MOV_HERRA"),
    "LOTES":              (lambda: os.path.join(REG, "Lotes.xlsx"),                  "LOTES"),
    "REG_AVANCE":         (lambda: os.path.join(REG, "RegAvance.xlsx"),              "REG_AVANCE"),
}

# Tablas que viven en users.db (SQLite), no en Excel
_SQLITE_TABLES = {"USUARIOS", "TALLERES", "AUDIT_LOG", "LOGIN_ATTEMPTS",
                  "DASHBOARD_CACHE", "ALERTAS", "ALERT_RULES"}


# ── Taller activo ─────────────────────────────────────────────────────────────

def _current_taller_id():
    """Devuelve el taller_id de la sesión activa, o None si es vista global."""
    tid = session.get("taller_id", "rober_lang")
    return None if tid == "__all__" else tid


# ── Helpers Excel ────────────────────────────────────────────────────────────

def _read_sheet(filepath, sheet=0):
    """Lee una hoja de Excel y devuelve DataFrame con strings."""
    df = pd.read_excel(filepath, sheet_name=sheet, dtype=str)
    return df.fillna("").astype(str)


def _df_to_list(df):
    return json.loads(df.to_json(orient="records", force_ascii=False))


def _append_row(filepath, sheet_name, row_dict):
    """Agrega una fila al final de la hoja indicada."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            abort(400, f"Hoja '{sheet_name}' no encontrada")
        ws = wb[sheet_name]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        new_row = [row_dict.get(h, "") for h in headers]
        ws.append(new_row)
        wb.save(filepath)
    except PermissionError:
        abort(500, f"Archivo bloqueado: {filepath}. Cierra Excel e intenta de nuevo.")
    except Exception as e:
        abort(500, f"Error al guardar: {e}")


def _update_cell(filepath, sheet_name, row_idx, col_name, value):
    """Actualiza una celda (row_idx 1-based en datos, sin contar header)."""
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet_name]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index(col_name) + 1
        ws.cell(row=row_idx + 1, column=col_idx, value=value)
        wb.save(filepath)
    except PermissionError:
        abort(500, f"Archivo bloqueado: {filepath}. Cierra Excel e intenta de nuevo.")
    except Exception as e:
        abort(500, f"Error al actualizar: {e}")


def _clear_sheet(filepath, sheet_name):
    """Borra todas las filas de datos manteniendo la cabecera."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        wb.save(filepath)
        return True
    except Exception:
        return False


def _ensure_columns(filepath, sheet_name, required_cols):
    """Añade columnas faltantes al final del header."""
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    existing = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    changed = False
    for col_name in required_cols:
        if col_name not in existing:
            ws.cell(row=1, column=len(existing) + 1, value=col_name)
            existing.append(col_name)
            changed = True
    if changed:
        wb.save(filepath)


# ── SQLite para USUARIOS ─────────────────────────────────────────────────────

def _users_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_users_db():
    """Crea la tabla USUARIOS y el usuario admin por defecto si no existe."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""CREATE TABLE IF NOT EXISTS USUARIOS (
        username TEXT, password_hash TEXT, nivel TEXT, nombre TEXT, activo TEXT, secciones TEXT
    )""")
    # Tablas auxiliares (vacías en local, necesarias para que admin no falle)
    conn.execute("""CREATE TABLE IF NOT EXISTS TALLERES (
        id TEXT PRIMARY KEY, nombre TEXT NOT NULL,
        activo INTEGER DEFAULT 1,
        creado_en TEXT DEFAULT (datetime('now')))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS AUDIT_LOG (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        taller_id TEXT, usuario TEXT, ip TEXT, accion TEXT,
        tabla TEXT, registro_id TEXT, antes TEXT, despues TEXT,
        ts TEXT DEFAULT (datetime('now')))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS LOGIN_ATTEMPTS (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ip TEXT, username TEXT, ts TEXT DEFAULT (datetime('now')))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS ALERTAS (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        taller_id TEXT, rule_id INTEGER, tipo TEXT,
        mensaje TEXT, referencia TEXT, leida INTEGER DEFAULT 0,
        ts TEXT DEFAULT (datetime('now')))""")
    conn.commit()
    # Migración: agregar columna secciones si falta
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(USUARIOS)").fetchall()]
        if "secciones" not in cols:
            conn.execute("ALTER TABLE USUARIOS ADD COLUMN secciones TEXT DEFAULT ''")
            conn.commit()
    except Exception:
        pass
    # Migración: rellenar secciones vacías según nivel
    try:
        for nivel, secs in DEFAULT_SECCIONES.items():
            conn.execute(
                "UPDATE USUARIOS SET secciones=? WHERE nivel=? AND (secciones IS NULL OR secciones='')",
                (secs, nivel)
            )
        conn.commit()
    except Exception:
        pass
    # Seed yekflowtotal
    try:
        conn.execute("""INSERT OR IGNORE INTO USUARIOS(username,password_hash,nivel,nombre,activo,secciones)
                        VALUES(?,?,?,?,?,?)""",
                     ("yekflowtotal", generate_password_hash("yekflow2025"),
                      "yekflow", "YekFlow Total", "1", DEFAULT_SECCIONES["yekflow"]))
        conn.commit()
    except Exception:
        pass
    # Unique index
    try:
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_usuarios_username ON USUARIOS(username)")
        conn.commit()
    except Exception:
        pass
    # Seed taller principal
    try:
        conn.execute("INSERT OR IGNORE INTO TALLERES(id, nombre) VALUES('rober_lang', 'Rober Lang')")
        conn.commit()
    except Exception:
        pass
    # Seed admin
    _db._col_cache.clear()
    rows = _db.table_to_list(conn, "USUARIOS")
    if not [r for r in rows if r.get("username") != "yekflowtotal"]:
        _db.insert_row(conn, "USUARIOS", {
            "username":      "admin",
            "password_hash": generate_password_hash("taller2024"),
            "nivel":         "superusuario",
            "nombre":        "Administrador",
            "activo":        "1",
            "secciones":     DEFAULT_SECCIONES["superusuario"],
        })
    conn.close()


# ── Interfaz pública (misma firma que sqlite_backend) ────────────────────────

def get_db():
    """Devuelve conexión a users.db (para auth y USUARIOS)."""
    return _users_conn()


def tlist(table):
    """Lista todos los registros de una tabla, filtrados por taller activo.
    Aislamiento estricto: solo devuelve datos del taller actual."""
    if table in _SQLITE_TABLES:
        conn = _users_conn()
        rows = _db.table_to_list(conn, table)
        conn.close()
        return rows
    if table not in _TABLE_MAP:
        return []
    fp_fn, sheet = _TABLE_MAP[table]
    try:
        df = _read_sheet(fp_fn(), sheet)
        rows = _df_to_list(df)
        # Filtrar por taller activo (aislamiento estricto)
        tid = _current_taller_id()
        if tid and table not in _CROSS_TALLER_TABLES and "taller_id" in df.columns:
            rows = [r for r in rows if r.get("taller_id") == tid]
        return rows
    except Exception:
        return []


def tfiltered(table, **kw):
    """Lista registros filtrados por columna=valor."""
    rows = tlist(table)
    return [r for r in rows if all(str(r.get(k, "")) == str(v) for k, v in kw.items())]


def tfiltered_dict(table, filters):
    """Lista registros filtrados por un dict {col: val}."""
    if table in _SQLITE_TABLES:
        conn = _users_conn()
        rows = _db.table_to_list_filtered(conn, table, filters)
        conn.close()
        return rows
    rows = tlist(table)
    return [r for r in rows if all(str(r.get(k, "")) == str(v) for k, v in filters.items())]


def tinsert(table, data):
    """Inserta una fila, inyectando taller_id del taller activo."""
    if table in _SQLITE_TABLES:
        conn = _users_conn()
        _db.insert_row(conn, table, data)
        conn.close()
        return
    if table not in _TABLE_MAP:
        return
    # Inyectar taller_id si no viene en data
    if table not in _CROSS_TALLER_TABLES and "taller_id" not in data:
        tid = _current_taller_id()
        data["taller_id"] = tid or "rober_lang"
    fp_fn, sheet = _TABLE_MAP[table]
    filepath = fp_fn()
    # Asegurar columnas necesarias
    if table == "MOV_ALMACEN":
        _ensure_columns(filepath, sheet, ["Hora", "Folio", "taller_id"])
    elif table == "MOV_HERRA":
        _ensure_columns(filepath, sheet, ["Hora", "Folio", "taller_id"])
    else:
        _ensure_columns(filepath, sheet, ["taller_id"])
    _append_row(filepath, sheet, data)


def tupdate_pk(table, pk_col, pk_val, data):
    """Actualiza una fila buscando por clave primaria."""
    if table in _SQLITE_TABLES:
        conn = _users_conn()
        result = _db.update_row_by_pk(conn, table, pk_col, pk_val, data)
        conn.close()
        return result
    if table not in _TABLE_MAP:
        return False
    fp_fn, sheet = _TABLE_MAP[table]
    filepath = fp_fn()
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if pk_col not in headers:
            return False
        pk_idx = headers.index(pk_col) + 1
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=pk_idx).value) == str(pk_val):
                for col, header in enumerate(headers, 1):
                    if header in data:
                        ws.cell(row=row, column=col, value=data[header])
                wb.save(filepath)
                return True
        return False
    except Exception:
        return False


def tdelete_rowid(table, rowid):
    """Elimina una fila por índice (1-based en datos)."""
    if table in _SQLITE_TABLES:
        conn = _users_conn()
        _db.delete_row_by_rowid(conn, table, rowid)
        conn.close()
        return
    if table not in _TABLE_MAP:
        return
    fp_fn, sheet = _TABLE_MAP[table]
    filepath = fp_fn()
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet]
        ws.delete_rows(int(rowid) + 1)  # +1 porque fila 1 = cabecera
        wb.save(filepath)
    except Exception:
        pass


def tclear(table):
    """Borra todas las filas de datos manteniendo la cabecera."""
    if table in _SQLITE_TABLES:
        conn = _users_conn()
        _db.clear_table(conn, table)
        conn.close()
        return True
    if table not in _TABLE_MAP:
        return False
    fp_fn, sheet = _TABLE_MAP[table]
    return _clear_sheet(fp_fn(), sheet)


# ── Raw table CRUD (para el editor de admin) ─────────────────────────────────

def raw_tabla_get(table_name):
    """Devuelve columnas y filas de una tabla para el editor CRUD."""
    if table_name in _SQLITE_TABLES:
        conn = _users_conn()
        cols = _db.get_columns(conn, table_name)
        rows = _db.table_to_list(conn, table_name)
        conn.close()
        return cols, rows
    # Buscar en _TABLE_MAP por nombre de tabla
    for key, (fp_fn, sheet) in _TABLE_MAP.items():
        if key == table_name:
            df = _read_sheet(fp_fn(), sheet)
            cols = list(df.columns)
            rows = []
            for i, (_, row) in enumerate(df.iterrows(), start=1):
                r = {"_idx": i}
                for c in cols:
                    val = row[c]
                    r[c] = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val)
                rows.append(r)
            return cols, rows
    return [], []


def raw_tabla_update(table_name, row_idx, col, value):
    """Actualiza una celda por índice de fila."""
    if table_name in _SQLITE_TABLES:
        conn = _users_conn()
        _db.update_cell_by_rowid(conn, table_name, int(row_idx), col, value)
        conn.close()
        return
    for key, (fp_fn, sheet) in _TABLE_MAP.items():
        if key == table_name:
            _update_cell(fp_fn(), sheet, int(row_idx), col, value)
            return


def raw_tabla_delete(table_name, row_idx):
    """Elimina una fila por índice."""
    if table_name in _SQLITE_TABLES:
        conn = _users_conn()
        _db.delete_row_by_rowid(conn, table_name, int(row_idx))
        conn.close()
        return
    for key, (fp_fn, sheet) in _TABLE_MAP.items():
        if key == table_name:
            filepath = fp_fn()
            wb = load_workbook(filepath)
            ws = wb[sheet]
            ws.delete_rows(int(row_idx) + 1)
            wb.save(filepath)
            return


# ── Stock ────────────────────────────────────────────────────────────────────

def calcular_stock():
    """Calcula stock actual por material: Σ Entradas − Σ Salidas."""
    mats = tlist("MATERIALES")
    movs = tlist("MOV_ALMACEN")
    result = []
    for m in mats:
        cod = m["Codigo_mat"]
        mat_movs = [mv for mv in movs if mv["Codigo_mat"] == cod]
        entradas = sum(safe_float(mv["Cantidad"]) for mv in mat_movs if mv["Tipo"] == "Entrada")
        salidas  = sum(safe_float(mv["Cantidad"]) for mv in mat_movs if mv["Tipo"] == "Salida")
        stock_actual = entradas - salidas
        stock_min = safe_float(m.get("Stock_min", 0))
        if stock_actual <= 0:
            nivel = "danger"
        elif stock_actual < stock_min:
            nivel = "warning"
        else:
            nivel = "ok"
        result.append({
            "Codigo_mat":   cod,
            "Descripcion":  m.get("Descripción", ""),
            "Unidad":       m.get("Unidad", ""),
            "Stock_min":    stock_min,
            "Stock_actual": round(float(stock_actual), 2),
            "Alerta":       nivel != "ok",
            "Nivel":        nivel,
        })
    return result


# ── Herramientas sync ────────────────────────────────────────────────────────

def _sync_herramienta_estado(id_herramienta):
    """Recalcula Estado y Responsable_Actual leyendo préstamos activos."""
    movs    = tfiltered("MOV_HERRA", ID_Herramienta=id_herramienta)
    activos = [m for m in movs if m.get("FechaDevolucion", "") == ""]
    if activos:
        nuevo_estado = "Prestada"
        responsable  = activos[-1].get("Responsable", "")
    else:
        nuevo_estado = "Disponible"
        responsable  = ""
    tupdate_pk("HERRAMIENTAS", "ID_Herramienta", id_herramienta, {
        "Estado": nuevo_estado,
        "Responsable_Actual": responsable,
    })


# ── Audit / Alertas (no-op o vacío para local) ──────────────────────────────

def audit_query(filters):
    """Devuelve entradas del AUDIT_LOG. Vacío en local."""
    conn = _users_conn()
    try:
        conds, params = [], []
        if filters.get("taller"):
            conds.append("taller_id = ?"); params.append(filters["taller"])
        if filters.get("tabla"):
            conds.append("tabla = ?");     params.append(filters["tabla"])
        if filters.get("usuario"):
            conds.append("usuario = ?");   params.append(filters["usuario"])
        if filters.get("desde"):
            conds.append("ts >= ?");       params.append(filters["desde"])
        if filters.get("hasta"):
            conds.append("ts <= ?");       params.append(filters["hasta"])
        where = f"WHERE {' AND '.join(conds)}" if conds else ""
        limit = filters.get("limit", 200)
        sql = f"SELECT * FROM AUDIT_LOG {where} ORDER BY ts DESC LIMIT ?"
        params.append(limit)
        rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    except Exception:
        rows = []
    conn.close()
    return rows


def alertas_query(taller_id, solo_no_leidas=True):
    """Devuelve alertas. Vacío en local (no se generan alertas)."""
    conn = _users_conn()
    try:
        cond = "WHERE leida=0" if solo_no_leidas else ""
        rows = [dict(r) for r in conn.execute(
            f"SELECT * FROM ALERTAS {cond} ORDER BY ts DESC LIMIT 200"
        ).fetchall()]
    except Exception:
        rows = []
    conn.close()
    return rows


def alerta_marcar_leida(aid):
    conn = _users_conn()
    try:
        conn.execute("UPDATE ALERTAS SET leida=1 WHERE id=?", (aid,))
        conn.commit()
    except Exception:
        pass
    conn.close()


def alertas_marcar_todas_leidas(taller_id):
    conn = _users_conn()
    try:
        conn.execute("UPDATE ALERTAS SET leida=1")
        conn.commit()
    except Exception:
        pass
    conn.close()


# ── Evaluar alertas (no-op para local) ───────────────────────────────────────

def _evaluar_alertas(conn, taller_id):
    pass


# ── Backup / Restore (ZIP de Excel) ──────────────────────────────────────────

def create_backup():
    """Genera un ZIP con todos los archivos Excel y lo devuelve como descarga."""
    import io
    import zipfile
    from flask import send_file
    mem = io.BytesIO()
    now_str = datetime.now().strftime("%Y-%m-%d_%H%M")
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for folder, prefix in [(CAT, "CAT"), (REG, "REG")]:
            if not os.path.isdir(folder):
                continue
            for fname in sorted(os.listdir(folder)):
                if fname.lower().endswith(".xlsx"):
                    fpath = os.path.join(folder, fname)
                    zf.write(fpath, f"{prefix}/{fname}")
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"backup_{now_str}.zip",
    )


def restore_backup(uploaded_file):
    """Restaura archivos Excel desde un ZIP de backup."""
    import zipfile
    from flask import jsonify
    if not uploaded_file.filename.lower().endswith(".zip"):
        return jsonify({"ok": False, "error": "El archivo debe ser un .zip"}), 400
    folder_map = {"CAT": CAT, "REG": REG}
    restored = []
    errors = []
    try:
        with zipfile.ZipFile(uploaded_file, "r") as zf:
            for name in zf.namelist():
                parts = name.replace("\\", "/").split("/")
                if len(parts) == 2 and parts[0] in folder_map and parts[1].lower().endswith(".xlsx") and parts[1]:
                    dest_dir = folder_map[parts[0]]
                    dest_path = os.path.join(dest_dir, parts[1])
                    try:
                        with zf.open(name) as src, open(dest_path, "wb") as dst:
                            dst.write(src.read())
                        restored.append(name)
                    except Exception as e:
                        errors.append(f"{name}: {str(e)}")
    except zipfile.BadZipFile:
        return jsonify({"ok": False, "error": "El ZIP es invalido o esta corrupto"}), 400
    if not restored:
        return jsonify({"ok": False, "error": "El ZIP no contiene archivos validos"}), 400
    return jsonify({"ok": True, "restored": restored, "errors": errors})


# ── Export deploy ────────────────────────────────────────────────────────────

def excel_to_sqlite_mem():
    """Migra todos los Excel a SQLite en memoria. Devuelve bytes del .db."""
    import tempfile
    import shutil
    tmp_dir = tempfile.mkdtemp()
    db_path = os.path.join(tmp_dir, "taller.db")
    try:
        _db.init_db(db_path)
        import sqlite3 as _sqlite3
        conn = _sqlite3.connect(db_path)
        for table, (fp_fn, sheet) in _TABLE_MAP.items():
            try:
                df = _read_sheet(fp_fn(), sheet)
                cols_excel = list(df.columns)
                cols_db = _db.get_columns(conn, table)
                common = [c for c in cols_excel if c in cols_db]
                if not common:
                    continue
                ph = ",".join("?" for _ in common)
                col_sql = ",".join(f'"{c}"' for c in common)
                sql = f'INSERT INTO "{table}" ({col_sql}) VALUES ({ph})'
                for _, row in df.iterrows():
                    vals = [str(row[c]) if str(row[c]) not in ("", "nan", "None") else None
                            for c in common]
                    conn.execute(sql, vals)
                conn.commit()
            except Exception:
                pass
        conn.close()
        with open(db_path, "rb") as f:
            return f.read()
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ── App wiring ───────────────────────────────────────────────────────────────

def init_app(app):
    """Inicializa la base de datos de usuarios."""
    _ensure_users_db()
