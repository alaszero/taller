
from openpyxl import load_workbook
import os

BASE = r'D:\Proyectos'
CAT  = os.path.join(BASE, 'CAT')
REG  = os.path.join(BASE, 'REG')

def clear_data(ws):
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)

def append_row(ws, row_dict, headers):
    ws.append([row_dict.get(h, '') for h in headers])

# ══════════════════════════════════════════════════════════════════
# 1. MATERIALES (55 materiales reales)
# ══════════════════════════════════════════════════════════════════
MATERIALES = [
    # (Codigo_mat, Tipo, Descripción, Unidad, Stock_min)
    # Del INVENTARIO (31)
    ('MAT001','Consumible','Cinta Masking tape Tuck 3/4','pza',2),
    ('MAT002','Herraje','Corredera Cierre Suave 400mm 30kg','pza',2),
    ('MAT003','Herraje','Corredera Cierre Suave 560mm 30kg','pza',2),
    ('MAT004','Consumible','Duretan Sellador Poliuretano','pza',1),
    ('MAT005','EPP','Lentes de seguridad Truper 14252','pza',1),
    ('MAT006','Consumible','Lija de agua grano 180','pza',3),
    ('MAT007','Consumible','Lija Orbital G120','pza',5),
    ('MAT008','Consumible','Lija Orbital G320','pza',5),
    ('MAT009','Consumible','Lija Orbital G60','pza',3),
    ('MAT010','Consumible','Lija para madera grano 60','pza',2),
    ('MAT011','Madera',"Madera Poplar 1\" 8' Fast",'pie',20),
    ('MAT012','Madera',"Madera Roble Blanco 1\" x 10'",'pie',5),
    ('MAT013','Consumible','Mastipren Uso Profesional','pza',1),
    ('MAT014','Tablero','MDF Natural 18mm 4x8','pza',2),
    ('MAT015','Tablero','MDF Natural 25mm 4x8','pza',1),
    ('MAT016','Tablero','MDF Natural 9mm 4x8','pza',2),
    ('MAT017','Tablero','MDF Nogal Desempatado 18mm 4x8','pza',1),
    ('MAT018','Consumible','No mas clavos Resistol','pza',1),
    ('MAT019','Consumible','No mas clavos Truper','pza',1),
    ('MAT020','Tablero','Okume 15mm 4x8','pza',1),
    ('MAT021','Consumible','Pen silicon banos y cocinas','pza',1),
    ('MAT022','Herraje','Pija Acero Inoxidable C8X3','pza',20),
    ('MAT023','Herraje','Pija Galvanizada 1/2','pza',50),
    ('MAT024','Herraje','Pija rapida 1 1/4"','pza',100),
    ('MAT025','Herraje','Pija rapida 2 1/2"','pza',50),
    ('MAT026','Herraje','Pija rapida calibre 10 3"','pza',50),
    ('MAT027','Herraje','Pivot Door Slide 8080PEZ 20','pza',1),
    ('MAT028','Consumible','Plastiprotector 10m2','pza',1),
    ('MAT029','Consumible','Sikasil Universal Silicona Multiproposito','pza',1),
    ('MAT030','Herraje','Sistema colgante para puerta corrediza DN80CF','pza',1),
    ('MAT031','Herraje','Taquete Fijacion 5/16','pza',30),
    # Del Reporte Diario - no estaban en INVENTARIO (24)
    ('MAT032','Consumible','Resistol 850','pza',1),
    ('MAT033','Herraje','Clavillo C-23 5/8','pza',50),
    ('MAT034','Herraje','Pija rapida 1 1/2"','pza',50),
    ('MAT035','Herraje','Pija rapida 2"','pza',50),
    ('MAT036','Consumible','Lija de agua grano 220','pza',3),
    ('MAT037','Consumible','Lija de agua grano 320','pza',3),
    ('MAT038','Consumible','Lija Orbital G220','pza',3),
    ('MAT039','Consumible','Lija para madera grano 80','pza',3),
    ('MAT040','Consumible','Cinta Masking tape tuck 3/4 Azul','pza',2),
    ('MAT041','Herraje','Bisagra Cuello 0 Hafele','pza',2),
    ('MAT042','Herraje','Bisagra Cuello 15 Hafele','pza',2),
    ('MAT043','Herraje','Corredera montaje inf 400mm Hafele','pza',2),
    ('MAT044','Herraje','Corredera montaje inf 500mm Hafele','pza',2),
    ('MAT045','Herraje','Clavo Galvanizado Calibre C18X5/8','pza',20),
    ('MAT046','Herraje','Clavo Galvanizado Calibre C18X3/4','pza',20),
    ('MAT047','Herraje','Clavo Galvanizado Calibre C18X1-1/2','pza',20),
    ('MAT048','Tablero','MDF Roble Rustico 6mm Enchapado 4x8','pza',1),
    ('MAT049','Tablero','MDF Natural 5mm 4x8','pza',1),
    ('MAT050','Madera',"Madera Nogal 1\" 12'",'pie',5),
    ('MAT051','Madera',"Madera Roble Blanco 1\" x 11'-12'",'pie',5),
    ('MAT052','Madera','Madera Tiras habilitadas 6X2','pza',5),
    ('MAT053','Herraje','Pija rapida 3"','pza',50),
    ('MAT054','Herraje','Pija Galvanizada 5/8','pza',50),
    ('MAT055','Herraje','Tope de piso Horde Media Luna','pza',1),
]

mat_hdrs = ['Codigo_mat','Tipo','Descripción','Unidad','Stock_min']
fp_mat = os.path.join(CAT, 'Materiales.xlsx')
wb_mat = load_workbook(fp_mat)
ws_mat = wb_mat['MATERIALES']
clear_data(ws_mat)
for m in MATERIALES:
    append_row(ws_mat, dict(zip(mat_hdrs, m)), mat_hdrs)
wb_mat.save(fp_mat)
print(f"✓ MATERIALES       → {len(MATERIALES)} materiales")

# ══════════════════════════════════════════════════════════════════
# 2. Proyectos.xlsx
# ══════════════════════════════════════════════════════════════════
fp_prj = os.path.join(CAT, 'Proyectos.xlsx')
wb_prj = load_workbook(fp_prj)

# — PROYECTOS —
PROYECTOS = [
    ('PROY001','Barrón','Casa Rancho','2026-02-12','2026-05-14','','En proceso'),
    ('PROY002','Abraham Barrón','Casa Guanajuato','2026-02-20','2026-04-30','','En proceso'),
]
prj_hdrs = ['ID_Proyecto','Cliente','Nombre_Obra','Inicio','Fin','Encargado','Estado']
ws_prj = wb_prj['PROYECTOS']
clear_data(ws_prj)
for p in PROYECTOS:
    append_row(ws_prj, dict(zip(prj_hdrs, p)), prj_hdrs)
print(f"✓ PROYECTOS        → {len(PROYECTOS)} proyectos")

# — MUEBLES —
RANCHO_MUEBLES = [
    'PUERTA DE INGRESO A COCINA DE SERVICIO',
    'PUERTA CORREDIZA DE INGRESO A ALACENA',
    'PUERTA DE INGRESO A MEDIO BANO VISITAS',
    'PUERTA DE INGRESO A RECAMARA 1',
    'PUERTA DE INGRESO A RECAMARA 2',
    'PUERTA DE INGRESO A VESTIDOR DE RECAMARA 1',
    'PUERTA DE INGRESO A VESTIDOR DE RECAMARA 2',
    'PUERTA DE INGRESO A BANO DE RECAMARA 1',
    'PUERTA DE INGRESO A BANO DE RECAMARA 2',
    'PUERTA DE INGRESO A RECAMARA PRINCIPAL',
    'PUERTARECAMARA PPAL A BANO',
    'PUERTA RECAMARAA PRINCIPAL A VESTIDOR',
    'MUEBLE DE BANO PRINCIPAL',
    'MUEBLE DE BANO VISITAS',
    'MUEBLE DE BANO DE RECAMARA 1',
    'MUEBLE DE BANO DE RECAMARA 2',
    'ESTRUCTURAS METALICAS PARA RECIBIR MUEBLES DE BANO',
    'VESTIDOR DE RECAMARA PRINCIPAL',
    'CLOSET DE BLANCOS',
    'RESPALDO Y FALDON EN VESTIDOR PRINCIPAL',
    'BASE SE CAMA DE RECAMARA PRINCIPAL',
    'BASE SE CAMA DE RECAMARA 1',
    'BASE SE CAMA DE RECAMARA 2',
    'PUERTA DE INGRESO PRINCIPAL',
    'VESTIDOR DE RECAMARA 1',
    'VESTIDOR DE RECAMARA 2',
    'RESPALDO Y FALDON EN VESTIDOR RECAMARA1',
    'RESPALDO Y FALDON EN VESTIDOR RECAMARA2',
    # Solo del Reporte Diario
    'General',
    'Mantenimiento',
    'Pendiente',
    'Pergolas Terraza',
    'Mueble TV planta baja',
]
GUAN_MUEBLES = ['Escritorio','Mesa','Mueble de Bano','Puerta','Closet']

MUEBLES = []
mue_name_to_id = {}

idx = 1
for nm in RANCHO_MUEBLES:
    mid = f'MUE{idx:03d}'
    mue_name_to_id[nm.upper()] = mid
    tipo = 'Closet' if 'CLOSET' in nm.upper() else \
           'Vestidor' if 'VESTIDOR' in nm.upper() else \
           'Mueble de Baño' if 'BANO' in nm.upper() or 'BAÑO' in nm.upper() else \
           'Cama' if 'CAMA' in nm.upper() else \
           'Puerta' if 'PUERTA' in nm.upper() else \
           'Estructura' if 'ESTRUCTURA' in nm.upper() else \
           'Mueble'
    MUEBLES.append({
        'ID_Mueble':   mid,
        'ID_Proyecto': 'PROY001',
        'Mueble':      nm,
        'Cantidad':    1,
        'Tipo':        tipo,
        'Area':        'Carpintería',
        'Observación': '',
    })
    idx += 1

for nm in GUAN_MUEBLES:
    mid = f'MUE{idx:03d}'
    mue_name_to_id[nm.upper()] = mid
    MUEBLES.append({
        'ID_Mueble':   mid,
        'ID_Proyecto': 'PROY002',
        'Mueble':      nm,
        'Cantidad':    1,
        'Tipo':        nm,
        'Area':        'Carpintería',
        'Observación': '',
    })
    idx += 1

# También mapear variaciones de nombre usadas en OPs de Casa Guanajuato
mue_name_to_id['ESCRITORIO']     = mue_name_to_id['ESCRITORIO']
mue_name_to_id['MESA']           = mue_name_to_id['MESA']
mue_name_to_id['MUEBLE DE BANO'] = mue_name_to_id['MUEBLE DE BANO']
mue_name_to_id['PUERTA']         = mue_name_to_id['PUERTA']
mue_name_to_id['CLOSET']         = mue_name_to_id['CLOSET']

mue_hdrs = ['ID_Mueble','ID_Proyecto','Mueble','Cantidad','Tipo','Area','Observación']
ws_mue = wb_prj['MUEBLES']
clear_data(ws_mue)
for m in MUEBLES:
    append_row(ws_mue, m, mue_hdrs)
print(f"✓ MUEBLES          → {len(MUEBLES)} muebles")

# — ORDENES DE PRODUCCION (109 OPs) —
OPS_RAW = [
    ('PROY001','OP001','PUERTA DE INGRESO A COCINA DE SERVICIO','Instalacion','Alejandro Flores','2026-02-12','2026-02-17'),
    ('PROY001','OP003','PUERTA CORREDIZA DE INGRESO A ALACENA','Instalacion','Alejandro Flores','2026-02-12','2026-02-17'),
    ('PROY001','OP005','PUERTA DE INGRESO A MEDIO BANO VISITAS','Instalacion','Alejandro Flores','2026-02-12','2026-02-17'),
    ('PROY001','OP013','PUERTA DE INGRESO A RECAMARA 1','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP015','PUERTA DE INGRESO A RECAMARA 2','En Proceso Fabricado','Gustavo','2026-02-12','2026-02-18'),
    ('PROY001','OP020','PUERTA DE INGRESO A VESTIDOR DE RECAMARA 1','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP022','PUERTA DE INGRESO A VESTIDOR DE RECAMARA 2','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP024','PUERTA DE INGRESO A BANO DE RECAMARA 1','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP026','PUERTA DE INGRESO A BANO DE RECAMARA 2','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP028','PUERTA DE INGRESO A RECAMARA PRINCIPAL','Terminado','','2026-02-12','2026-02-12'),
    ('PROY001','OP029','PUERTARECAMARA PPAL A BANO','En Proceso Fabricado','Alejandro Flores','2026-02-12','2026-02-18'),
    ('PROY001','OP033','PUERTA RECAMARAA PRINCIPAL A VESTIDOR','En Proceso Fabricado','Alejandro Flores','2026-02-12','2026-02-18'),
    ('PROY001','OP037','MUEBLE DE BANO PRINCIPAL','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP039','MUEBLE DE BANO VISITAS','Laqueado','Jesus Valdez','2026-02-12','2026-02-18'),
    ('PROY001','OP043','MUEBLE DE BANO DE RECAMARA 1','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP045','MUEBLE DE BANO DE RECAMARA 2','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP047','ESTRUCTURAS METALICAS PARA RECIBIR MUEBLES DE BANO','Instalacion','','2026-02-12','2026-02-16'),
    ('PROY001','OP049','VESTIDOR DE RECAMARA PRINCIPAL','En Proceso Fabricado','Luis Tejeda','2026-02-12','2026-03-14'),
    ('PROY001','OP054','CLOSET DE BLANCOS','En Proceso Fabricado','Federico Munoz','2026-02-12','2026-02-25'),
    ('PROY001','OP063','RESPALDO Y FALDON EN VESTIDOR PRINCIPAL','En Proceso Fabricado','','2026-02-12','2026-03-14'),
    ('PROY001','OP130','BASE SE CAMA DE RECAMARA PRINCIPAL','En Proceso Fabricado','','2026-02-12','2026-02-28'),
    ('PROY001','OP135','BASE SE CAMA DE RECAMARA 1','En Proceso Fabricado','','2026-02-12','2026-02-28'),
    ('PROY001','OP140','BASE SE CAMA DE RECAMARA 2','En Proceso Fabricado','','2026-02-12','2026-02-28'),
    ('PROY001','OP007','PUERTA DE INGRESO PRINCIPAL','Fabricacion De Estructura En PTR','','2026-02-16','2026-02-21'),
    ('PROY001','OP002','PUERTA DE INGRESO A COCINA DE SERVICIO','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP004','PUERTA CORREDIZA DE INGRESO A ALACENA','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP006','PUERTA DE INGRESO A MEDIO BANO VISITAS','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP014','PUERTA DE INGRESO A RECAMARA 1','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP021','PUERTA DE INGRESO A VESTIDOR DE RECAMARA 1','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP023','PUERTA DE INGRESO A VESTIDOR DE RECAMARA 2','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP025','PUERTA DE INGRESO A BANO DE RECAMARA 1','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP027','PUERTA DE INGRESO A BANO DE RECAMARA 2','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP038','MUEBLE DE BANO PRINCIPAL','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP044','MUEBLE DE BANO DE RECAMARA 1','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP046','MUEBLE DE BANO DE RECAMARA 2','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP048','ESTRUCTURAS METALICAS PARA RECIBIR MUEBLES DE BANO','Terminado','','2026-02-17','2026-02-17'),
    ('PROY001','OP059','VESTIDOR DE RECAMARA 1','Instalacion','','2026-02-18','2026-02-21'),
    ('PROY001','OP061','VESTIDOR DE RECAMARA 2','Instalacion','','2026-02-18','2026-02-21'),
    ('PROY001','OP068','RESPALDO Y FALDON EN VESTIDOR RECAMARA1','Instalacion','','2026-02-18','2026-02-21'),
    ('PROY001','OP070','RESPALDO Y FALDON EN VESTIDOR RECAMARA2','Instalacion','','2026-02-18','2026-02-21'),
    ('PROY001','OP016','PUERTA DE INGRESO A RECAMARA 2','Laqueado','Gustavo','2026-02-19','2026-02-21'),
    ('PROY001','OP030','PUERTARECAMARA PPAL A BANO','Laqueado','','2026-02-19','2026-02-21'),
    ('PROY001','OP034','PUERTA RECAMARAA PRINCIPAL A VESTIDOR','Laqueado','','2026-02-19','2026-02-21'),
    ('PROY002','OP145','ESCRITORIO','En Proceso Fabricado','Abraham Barron','2026-02-20','2026-03-02'),
    ('PROY002','OP149','MESA','En Proceso Fabricado','Abraham Barron','2026-02-20','2026-03-02'),
    ('PROY002','OP153','MUEBLE DE BANO','En Proceso Fabricado','Abraham Barron','2026-02-20','2026-03-17'),
    ('PROY002','OP158','PUERTA','En Proceso Fabricado','Abraham Barron','2026-02-20','2026-02-23'),
    ('PROY002','OP163','CLOSET','En Proceso Fabricado','Abraham Barron','2026-02-20','2026-04-11'),
    ('PROY001','OP008','PUERTA DE INGRESO PRINCIPAL','En Proceso Fabricado','','2026-02-23','2026-03-14'),
    ('PROY001','OP031','PUERTARECAMARA PPAL A BANO','Instalacion','','2026-02-23','2026-02-23'),
    ('PROY001','OP040','MUEBLE DE BANO VISITAS','Instalacion','','2026-02-23','2026-02-24'),
    ('PROY001','OP060','VESTIDOR DE RECAMARA 1','Terminado','','2026-02-23','2026-02-23'),
    ('PROY001','OP062','VESTIDOR DE RECAMARA 2','Terminado','','2026-02-23','2026-02-23'),
    ('PROY001','OP069','RESPALDO Y FALDON EN VESTIDOR RECAMARA1','Terminado','','2026-02-23','2026-02-23'),
    ('PROY001','OP071','RESPALDO Y FALDON EN VESTIDOR RECAMARA2','Terminado','','2026-02-23','2026-02-23'),
    ('PROY002','OP159','PUERTA','Laqueado','Abraham Barron','2026-02-23','2026-03-03'),
    ('PROY001','OP032','PUERTARECAMARA PPAL A BANO','Terminado','','2026-02-24','2026-02-24'),
    ('PROY001','OP035','PUERTA RECAMARAA PRINCIPAL A VESTIDOR','Instalacion','','2026-02-24','2026-02-24'),
    ('PROY001','OP017','PUERTA DE INGRESO A RECAMARA 2','Instalacion','Gustavo','2026-02-25','2026-02-25'),
    ('PROY001','OP036','PUERTA RECAMARAA PRINCIPAL A VESTIDOR','Terminado','','2026-02-25','2026-02-25'),
    ('PROY001','OP055','CLOSET DE BLANCOS','Laqueado','','2026-02-26','2026-03-05'),
    ('PROY001','OP131','BASE SE CAMA DE RECAMARA PRINCIPAL','Laqueado','','2026-03-02','2026-03-14'),
    ('PROY001','OP136','BASE SE CAMA DE RECAMARA 1','Laqueado','','2026-03-02','2026-03-14'),
    ('PROY001','OP141','BASE SE CAMA DE RECAMARA 2','Laqueado','','2026-03-02','2026-03-14'),
    ('PROY002','OP146','ESCRITORIO','Laqueado','Abraham Barron','2026-03-02','2026-03-05'),
    ('PROY002','OP150','MESA','Laqueado','Abraham Barron','2026-03-02','2026-03-10'),
    ('PROY002','OP160','PUERTA','Instalacion','Abraham Barron','2026-03-04','2026-03-08'),
    ('PROY002','OP147','ESCRITORIO','Instalacion','Abraham Barron','2026-03-05','2026-03-12'),
    ('PROY001','OP056','CLOSET DE BLANCOS','Instalacion','','2026-03-06','2026-03-13'),
    ('PROY002','OP161','PUERTA','Detallado en Obra','Abraham Barron','2026-03-08','2026-03-10'),
    ('PROY001','OP018','PUERTA DE INGRESO A RECAMARA 2','Detallado en Obra','Gustavo','2026-03-09','2026-03-21'),
    ('PROY001','OP041','MUEBLE DE BANO VISITAS','Detallado en Obra','','2026-03-09','2026-03-21'),
    ('PROY001','OP019','PUERTA DE INGRESO A RECAMARA 2','Terminado','Gustavo','2026-03-10','2026-03-10'),
    ('PROY002','OP151','MESA','Instalacion','Abraham Barron','2026-03-10','2026-03-12'),
    ('PROY002','OP162','PUERTA','Terminado','Abraham Barron','2026-03-10','2026-03-10'),
    ('PROY002','OP148','ESCRITORIO','Terminado','Abraham Barron','2026-03-12','2026-03-12'),
    ('PROY002','OP152','MESA','Terminado','Abraham Barron','2026-03-12','2026-03-12'),
    ('PROY001','OP057','CLOSET DE BLANCOS','Detallado en Obra','','2026-03-14','2026-03-25'),
    ('PROY001','OP009','PUERTA DE INGRESO PRINCIPAL','Proceso Especial (aceitado)','','2026-03-16','2026-03-21'),
    ('PROY001','OP050','VESTIDOR DE RECAMARA PRINCIPAL','Laqueado','','2026-03-16','2026-03-28'),
    ('PROY001','OP064','RESPALDO Y FALDON EN VESTIDOR PRINCIPAL','Laqueado','','2026-03-16','2026-03-28'),
    ('PROY001','OP132','BASE SE CAMA DE RECAMARA PRINCIPAL','Instalacion','','2026-03-17','2026-03-21'),
    ('PROY001','OP137','BASE SE CAMA DE RECAMARA 1','Instalacion','','2026-03-17','2026-03-21'),
    ('PROY001','OP142','BASE SE CAMA DE RECAMARA 2','Instalacion','','2026-03-17','2026-03-21'),
    ('PROY002','OP154','MUEBLE DE BANO','Laqueado','Abraham Barron','2026-03-17','2026-03-22'),
    ('PROY002','OP155','MUEBLE DE BANO','Detallado','Abraham Barron','2026-03-22','2026-03-24'),
    ('PROY001','OP010','PUERTA DE INGRESO PRINCIPAL','Instalacion','','2026-03-23','2026-03-28'),
    ('PROY001','OP042','MUEBLE DE BANO VISITAS','Terminado','','2026-03-23','2026-03-23'),
    ('PROY001','OP133','BASE SE CAMA DE RECAMARA PRINCIPAL','Detallado en Obra','','2026-03-23','2026-03-24'),
    ('PROY001','OP134','BASE SE CAMA DE RECAMARA PRINCIPAL','Terminado','','2026-03-25','2026-03-25'),
    ('PROY001','OP138','BASE SE CAMA DE RECAMARA 1','Detallado en Obra','','2026-03-25','2026-03-26'),
    ('PROY002','OP156','MUEBLE DE BANO','Instalacion','Abraham Barron','2026-03-25','2026-03-28'),
    ('PROY001','OP058','CLOSET DE BLANCOS','Terminado','','2026-03-26','2026-03-26'),
    ('PROY001','OP139','BASE SE CAMA DE RECAMARA 1','Terminado','','2026-03-27','2026-03-27'),
    ('PROY001','OP143','BASE SE CAMA DE RECAMARA 2','Detallado en Obra','','2026-03-27','2026-03-27'),
    ('PROY001','OP144','BASE SE CAMA DE RECAMARA 2','Terminado','','2026-03-28','2026-03-28'),
    ('PROY002','OP157','MUEBLE DE BANO','Terminado','Abraham Barron','2026-03-28','2026-03-28'),
    ('PROY001','OP011','PUERTA DE INGRESO PRINCIPAL','Detallado en Obra','','2026-03-30','2026-04-02'),
    ('PROY001','OP051','VESTIDOR DE RECAMARA PRINCIPAL','Instalacion','','2026-03-30','2026-04-18'),
    ('PROY001','OP012','PUERTA DE INGRESO PRINCIPAL','Terminado','','2026-04-03','2026-04-03'),
    ('PROY002','OP164','CLOSET','Laqueado','Abraham Barron','2026-04-11','2026-04-16'),
    ('PROY002','OP165','CLOSET','Instalacion','Abraham Barron','2026-04-17','2026-04-27'),
    ('PROY001','OP052','VESTIDOR DE RECAMARA PRINCIPAL','Detallado en Obra','','2026-04-20','2026-05-09'),
    ('PROY001','OP065','RESPALDO Y FALDON EN VESTIDOR PRINCIPAL','Detallado','','2026-04-20','2026-05-09'),
    ('PROY002','OP166','CLOSET','Detallado en Obra','Abraham Barron','2026-04-27','2026-04-29'),
    ('PROY002','OP167','CLOSET','Terminado','Abraham Barron','2026-04-30','2026-04-30'),
    ('PROY001','OP053','VESTIDOR DE RECAMARA PRINCIPAL','Terminado','','2026-05-11','2026-05-11'),
    ('PROY001','OP066','RESPALDO Y FALDON EN VESTIDOR PRINCIPAL','Instalacion','','2026-05-11','2026-05-13'),
    ('PROY001','OP067','RESPALDO Y FALDON EN VESTIDOR PRINCIPAL','Terminado','','2026-05-14','2026-05-14'),
]

op_hdrs = ['ID_OP','ID_Mueble','Etapa','Cantidad','FechaInicioPlaneada','FechaFinalPlaneada','FechaInicioReal','FechaFinalReal','Responsable']
ws_op = wb_prj['ORDENES_PRODUCCION']
clear_data(ws_op)
sin_mueble = []
for tup in OPS_RAW:
    proy, op_id, mue_name, etapa, resp, ini, fin = tup
    mue_id = mue_name_to_id.get(mue_name.upper())
    if not mue_id:
        sin_mueble.append((op_id, mue_name))
        mue_id = mue_name  # fallback
    append_row(ws_op, {
        'ID_OP': op_id, 'ID_Mueble': mue_id, 'Etapa': etapa,
        'Cantidad': 1, 'FechaInicioPlaneada': ini, 'FechaFinalPlaneada': fin,
        'FechaInicioReal': '', 'FechaFinalReal': '', 'Responsable': resp,
    }, op_hdrs)
print(f"✓ ORDENES          → {len(OPS_RAW)} OPs")
if sin_mueble:
    print(f"  ⚠ Sin mueble mapeado: {sin_mueble}")

# — ETAPAS: agregar las nuevas que no existan —
ws_et = wb_prj['ETAPAS']
existentes = set()
for r in ws_et.iter_rows(min_row=2, values_only=True):
    if r[0]: existentes.add(str(r[0]))

NUEVAS_ETAPAS = [
    ('En Proceso Fabricado',        11, 'Carpintería'),
    ('Laqueado',                    12, 'Lacado'),
    ('Instalacion',                 13, 'Carpintería'),
    ('Detallado en Obra',           14, 'Carpintería'),
    ('Detallado',                   15, 'Lacado'),
    ('Terminado',                   16, 'Carpintería'),
    ('Fabricacion De Estructura En PTR', 17, 'Carpintería'),
    ('Proceso Especial (aceitado)', 18, 'Lacado'),
]
et_hdrs = ['Etapa','Orden','Area']
added = 0
for et in NUEVAS_ETAPAS:
    if et[0] not in existentes:
        append_row(ws_et, dict(zip(et_hdrs, et)), et_hdrs)
        added += 1
print(f"✓ ETAPAS           → {added} etapas nuevas")

wb_prj.save(fp_prj)

# ══════════════════════════════════════════════════════════════════
# 3. MOV_ALMACEN — SOLO Entradas iniciales = inventario actual
# ══════════════════════════════════════════════════════════════════
STOCK_INICIAL = {
    'MAT001': 4,   'MAT002': 4,   'MAT003': 5,   'MAT004': 2,
    'MAT005': 3,   'MAT006': 9,   'MAT007': 14,  'MAT008': 33,
    'MAT009': 7,   'MAT010': 5,   'MAT011': 90,
    'MAT013': 4,   'MAT014': 3,   'MAT015': 1,   'MAT016': 5,
    'MAT017': 1,   'MAT018': 2,   'MAT019': 3,   'MAT020': 2,
    'MAT021': 2,   'MAT022': 59,  'MAT023': 600, 'MAT024': 1000,
    'MAT025': 295, 'MAT026': 300,
    'MAT029': 2,   'MAT030': 2,   'MAT031': 150,
}
UNIDADES = {m[0]: m[3] for m in MATERIALES}

mov_hdrs = ['Fecha','Codigo_mat','Tipo','Cantidad','Unidad','ID_Proyecto','Mueble','ID_Trab','ID_Ubic_Origen','ID_Ubic_Destino']
fp_mov = os.path.join(REG, 'RegistroMovMaterial.xlsx')
wb_mov = load_workbook(fp_mov)
ws_mov = wb_mov['MOV_ALMACEN']
clear_data(ws_mov)

count = 0
for cod, qty in STOCK_INICIAL.items():
    if qty > 0:
        append_row(ws_mov, {
            'Fecha': '2026-01-01', 'Codigo_mat': cod, 'Tipo': 'Entrada',
            'Cantidad': qty, 'Unidad': UNIDADES.get(cod,'pza'),
            'ID_Proyecto': '', 'Mueble': 'Inventario inicial',
            'ID_Trab': '', 'ID_Ubic_Origen': '', 'ID_Ubic_Destino': '',
        }, mov_hdrs)
        count += 1

wb_mov.save(fp_mov)
print(f"✓ MOV_ALMACEN      → {count} entradas iniciales (stock correcto, sin negativos)")
print(f"\n✅ Migración completa.")
print(f"   Materiales: {len(MATERIALES)} | Muebles: {len(MUEBLES)} | OPs: {len(OPS_RAW)} | Stock inicial: {count} movimientos")
