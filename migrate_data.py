from openpyxl import load_workbook
import os

BASE = r'D:\Proyectos'
CAT  = os.path.join(BASE, 'CAT')
REG  = os.path.join(BASE, 'REG')

# ══════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════
def clear_data(ws):
    """Elimina todas las filas de datos (mantiene encabezado fila 1)."""
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)

def append(ws, row_dict, headers):
    """Agrega una fila según el orden de los encabezados."""
    ws.append([row_dict.get(h, '') for h in headers])

# ══════════════════════════════════════════════════════════════════
# 1. MATERIALES.xlsx  – hoja MATERIALES
# ══════════════════════════════════════════════════════════════════
# 31 materiales del INVENTARIO + 9 del Reporte Diario
MATERIALES = [
    # (Codigo_mat, Tipo, Descripción, Unidad, Stock_min)
    ('MAT001','Consumible','Cinta Masking tape Tuck 3/4','pza',2),
    ('MAT002','Herraje','Corredera Cierre Suave 400mm 30kg','pza',2),
    ('MAT003','Herraje','Corredera Cierre Suave 560mm 30kg','pza',2),
    ('MAT004','Consumible','Duretan Sellador Poliuretano','pza',1),
    ('MAT005','EPP','Lentes de seguridad Truper 14252','pza',1),
    ('MAT006','Consumible','Lija de agua grano 180','pza',5),
    ('MAT007','Consumible','Lija Orbital G120','pza',5),
    ('MAT008','Consumible','Lija Orbital G320','pza',10),
    ('MAT009','Consumible','Lija Orbital G60','pza',3),
    ('MAT010','Consumible','Lija para madera grano 60','pza',3),
    ('MAT011','Madera',"Madera Poplar 1\" 8' Fast",'pie',30),
    ('MAT012','Madera',"Madera Roble Blanco 1\" x 10'",'pie',10),
    ('MAT013','Consumible','Mastipren Uso Profesional','pza',1),
    ('MAT014','Tablero','MDF Natural 18mm 4x8','pza',2),
    ('MAT015','Tablero','MDF Natural 25mm 4x8','pza',1),
    ('MAT016','Tablero','MDF Natural 9mm 4x8','pza',2),
    ('MAT017','Tablero','MDF Nogal Desempatado 18mm 4x8','pza',1),
    ('MAT018','Consumible','No mas clavos Resistol','pza',1),
    ('MAT019','Consumible','No mas clavos Truper','pza',1),
    ('MAT020','Tablero','Okume 15mm 4x8','pza',1),
    ('MAT021','Consumible','Pen silicon banos y cocinas','pza',1),
    ('MAT022','Herraje','Pija Acero Inoxidable C8X3','pza',20),
    ('MAT023','Herraje','Pija Galvanizada 1/2','pza',100),
    ('MAT024','Herraje','Pija rapida 1 1/4"','pza',100),
    ('MAT025','Herraje','Pija rapida 2 1/2"','pza',50),
    ('MAT026','Herraje','Pija rapida calibre 10 3"','pza',50),
    ('MAT027','Herraje','Pivot Door Slide 8080PEZ 20','pza',1),
    ('MAT028','Consumible','Plastiprotector 10m2','pza',1),
    ('MAT029','Consumible','Sikasil Universal Silicona Multiproposito','pza',1),
    ('MAT030','Herraje','Sistema colgante para puerta corrediza DN80CF','pza',1),
    ('MAT031','Herraje','Taquete Fijacion 5/16','pza',30),
    # Adicionales del Reporte Diario (no estaban en INVENTARIO)
    ('MAT032','Consumible','Resistol 850','pza',1),
    ('MAT033','Herraje','Clavillo C-23 5/8','pza',50),
    ('MAT034','Herraje','Pija rapida 1 1/2"','pza',100),
    ('MAT035','Herraje','Pija rapida 2"','pza',100),
    ('MAT036','Consumible','Lija de agua grano 220','pza',3),
    ('MAT037','Consumible','Lija de agua grano 320','pza',3),
    ('MAT038','Consumible','Lija Orbital G220','pza',5),
    ('MAT039','Consumible','Lija para madera grano 80','pza',3),
    ('MAT040','Consumible','Cinta Masking tape tuck 3/4 Azul','pza',2),
]

mat_hdrs = ['Codigo_mat','Tipo','Descripción','Unidad','Stock_min']
fp = os.path.join(CAT, 'Materiales.xlsx')
wb = load_workbook(fp)
ws = wb['MATERIALES']
clear_data(ws)
for m in MATERIALES:
    append(ws, dict(zip(mat_hdrs, m)), mat_hdrs)
wb.save(fp)
print(f"✓ Materiales.xlsx  → {len(MATERIALES)} materiales")

# ══════════════════════════════════════════════════════════════════
# 2. Proyectos.xlsx
# ══════════════════════════════════════════════════════════════════
fp2 = os.path.join(CAT, 'Proyectos.xlsx')
wb2 = load_workbook(fp2)

# — PROYECTOS —
PROYECTOS = [
    ('PROY001','Abraham Barrón','Casa Guanajuato','2026-02-20','2026-04-30','','En proceso'),
    ('PROY002','(Por definir)','Proyecto Aqua','2026-01-01','','','En proceso'),
]
prj_hdrs = ['ID_Proyecto','Cliente','Nombre_Obra','Inicio','Fin','Encargado','Estado']
ws_prj = wb2['PROYECTOS']
clear_data(ws_prj)
for p in PROYECTOS:
    append(ws_prj, dict(zip(prj_hdrs, p)), prj_hdrs)
print(f"✓ PROYECTOS        → {len(PROYECTOS)} proyectos")

# — MUEBLES —
MUEBLES = [
    # Casa Guanajuato
    ('MUE001','PROY001','Escritorio',1,'Escritorio','Carpintería',''),
    ('MUE002','PROY001','Mesa',1,'Mesa','Carpintería',''),
    ('MUE003','PROY001','Mueble de Baño',1,'Mueble de Baño','Carpintería',''),
    ('MUE004','PROY001','Puerta',1,'Puerta','Carpintería',''),
    ('MUE005','PROY001','Closet',1,'Closet','Carpintería',''),
    # Proyecto Aqua
    ('MUE006','PROY002','Closet de Blancos',1,'Closet','Carpintería',''),
    ('MUE007','PROY002','Credeza',1,'Credenza','Carpintería',''),
    ('MUE008','PROY002','Puerta Aqua General',1,'Puerta','Carpintería',''),
    ('MUE009','PROY002','Puerta Corrediza',1,'Puerta','Carpintería',''),
    ('MUE010','PROY002','Puerta Escalera Aqua',1,'Puerta','Lacado',''),
    ('MUE011','PROY002','Vestidor Recámara Principal',1,'Vestidor','Carpintería',''),
]
mue_hdrs = ['ID_Mueble','ID_Proyecto','Mueble','Cantidad','Tipo','Area','Observación']
ws_mue = wb2['MUEBLES']
clear_data(ws_mue)
for m in MUEBLES:
    append(ws_mue, dict(zip(mue_hdrs, m)), mue_hdrs)
print(f"✓ MUEBLES          → {len(MUEBLES)} muebles")

# — ORDENES_PRODUCCION —
# Mapa nombre → ID_Mueble (Casa Guanajuato)
mue_map = {
    'ESCRITORIO':'MUE001','MESA':'MUE002','MUEBLE DE BANO':'MUE003',
    'PUERTA':'MUE004','CLOSET':'MUE005',
}
OPS_RAW = [
    ('OP145','ESCRITORIO','En Proceso Fabricado','2026-02-20','2026-03-02'),
    ('OP149','MESA','En Proceso Fabricado','2026-02-20','2026-03-02'),
    ('OP153','MUEBLE DE BANO','En Proceso Fabricado','2026-02-20','2026-03-17'),
    ('OP158','PUERTA','En Proceso Fabricado','2026-02-20','2026-02-23'),
    ('OP163','CLOSET','En Proceso Fabricado','2026-02-20','2026-04-11'),
    ('OP159','PUERTA','Laqueado','2026-02-23','2026-03-03'),
    ('OP146','ESCRITORIO','Laqueado','2026-03-02','2026-03-05'),
    ('OP150','MESA','Laqueado','2026-03-02','2026-03-10'),
    ('OP160','PUERTA','Instalacion','2026-03-04','2026-03-08'),
    ('OP147','ESCRITORIO','Instalacion','2026-03-05','2026-03-12'),
    ('OP161','PUERTA','Detallado en Obra','2026-03-08','2026-03-10'),
    ('OP151','MESA','Instalacion','2026-03-10','2026-03-12'),
    ('OP162','PUERTA','Terminado','2026-03-10','2026-03-10'),
    ('OP148','ESCRITORIO','Terminado','2026-03-12','2026-03-12'),
    ('OP152','MESA','Terminado','2026-03-12','2026-03-12'),
    ('OP154','MUEBLE DE BANO','Laqueado','2026-03-17','2026-03-22'),
    ('OP155','MUEBLE DE BANO','Detallado','2026-03-22','2026-03-24'),
    ('OP156','MUEBLE DE BANO','Instalacion','2026-03-25','2026-03-28'),
    ('OP157','MUEBLE DE BANO','Terminado','2026-03-28','2026-03-28'),
    ('OP164','CLOSET','Laqueado','2026-04-11','2026-04-16'),
    ('OP165','CLOSET','Instalacion','2026-04-17','2026-04-27'),
    ('OP166','CLOSET','Detallado en Obra','2026-04-27','2026-04-29'),
    ('OP167','CLOSET','Terminado','2026-04-30','2026-04-30'),
]
ORDENES = []
for op_id, mue_name, etapa, ini, fin in OPS_RAW:
    ORDENES.append({
        'ID_OP': op_id,
        'ID_Mueble': mue_map.get(mue_name, mue_name),
        'Etapa': etapa,
        'Cantidad': 1,
        'FechaInicioPlaneada': ini,
        'FechaFinalPlaneada': fin,
        'FechaInicioReal': '',
        'FechaFinalReal': '',
        'Responsable': '',
    })
op_hdrs = ['ID_OP','ID_Mueble','Etapa','Cantidad','FechaInicioPlaneada','FechaFinalPlaneada','FechaInicioReal','FechaFinalReal','Responsable']
ws_op = wb2['ORDENES_PRODUCCION']
clear_data(ws_op)
for o in ORDENES:
    append(ws_op, o, op_hdrs)
print(f"✓ ORDENES          → {len(ORDENES)} OPs")

# — ETAPAS (agregar etapas nuevas que vienen de las OPs reales) —
ws_et = wb2['ETAPAS']
etapas_existentes = set()
for row in ws_et.iter_rows(min_row=2, values_only=True):
    if row[0]:
        etapas_existentes.add(str(row[0]))

nuevas_etapas = [
    ('Fabricado',           11, 'Carpintería'),
    ('Laqueado',            12, 'Lacado'),
    ('Instalacion',         13, 'Carpintería'),
    ('Detallado en Obra',   14, 'Carpintería'),
    ('Detallado',           15, 'Lacado'),
    ('Terminado',           16, 'Carpintería'),
]
et_hdrs = ['Etapa','Orden','Area']
added_et = 0
for et in nuevas_etapas:
    if et[0] not in etapas_existentes:
        append(ws_et, dict(zip(et_hdrs, et)), et_hdrs)
        added_et += 1
print(f"✓ ETAPAS           → {added_et} etapas nuevas agregadas")

wb2.save(fp2)

# ══════════════════════════════════════════════════════════════════
# 3. RegistroMovMaterial.xlsx  – hoja MOV_ALMACEN
# ══════════════════════════════════════════════════════════════════
# Inventario inicial por material (Codigo_mat, cantidad_inicial)
STOCK_INICIAL = {
    'MAT001': 4,   # Cinta Masking tape Tuck 3/4
    'MAT002': 4,   # Corredera 400mm
    'MAT003': 5,   # Corredera 560mm
    'MAT004': 2,   # Duretan
    'MAT005': 3,   # Lentes
    'MAT006': 9,   # Lija agua 180
    'MAT007': 14,  # Lija Orbital G120
    'MAT008': 33,  # Lija Orbital G320
    'MAT009': 7,   # Lija Orbital G60
    'MAT010': 5,   # Lija madera 60
    'MAT011': 90,  # Madera Poplar
    # MAT012 Roble: sin inventario
    'MAT013': 4,   # Mastipren
    'MAT014': 3,   # MDF 18mm
    'MAT015': 1,   # MDF 25mm
    'MAT016': 5,   # MDF 9mm
    'MAT017': 1,   # MDF Nogal
    'MAT018': 2,   # No mas clavos Resistol
    'MAT019': 3,   # No mas clavos Truper
    'MAT020': 2,   # Okume 15mm
    'MAT021': 2,   # Pen silicon
    'MAT022': 59,  # Pija Acero Inoxidable
    'MAT023': 600, # Pija Galvanizada
    'MAT024': 1000,# Pija rapida 1 1/4
    'MAT025': 295, # Pija rapida 2 1/2
    'MAT026': 300, # Pija rapida calibre 10 3"
    # MAT027 Pivot: 0 (agotado)
    # MAT028 Plastiprotector: 0
    'MAT029': 2,   # Sikasil
    'MAT030': 2,   # Sistema colgante
    'MAT031': 150, # Taquete
}

# Unidades por material
UNIDADES = {m[0]: m[3] for m in MATERIALES}

# Mapa nombre (Reporte Diario) → Codigo_mat
DESC_TO_CODE = {
    'Resistol 850':                              'MAT032',
    'Clavillo C-23 - 5/8':                       'MAT033',
    'Pivot Door Slide 8080PEZ 20':               'MAT027',
    'Pija rapida 1 1/2 "':                       'MAT034',
    'Pija rapida 1 1/4 "':                       'MAT024',
    'Pija rapida 2 1/2 "':                       'MAT025',
    'Pija rapida 2"':                            'MAT035',
    'Pija rapida calibre 10 3"':                 'MAT026',
    'Sistema colgante para puerta corrediza DN80CF': 'MAT030',
    'Cinta Masking tape Tuck 3/4':               'MAT001',
    'Cinta Masking tape tuck 3/4 Azul':          'MAT040',
    'Lija de agua grano 220':                    'MAT036',
    'Lija de agua grano 320':                    'MAT037',
    'Lija Orbital G120':                         'MAT007',
    'Lija Orbital G220':                         'MAT038',
    'Lija para madera grano 80':                 'MAT039',
}

# Mapa nombre mueble (Reporte Diario) → ID_Mueble
MUE_NAME_MAP = {
    'CLOSET DE BLANCOS':          'MUE006',
    'Credeza':                    'MUE007',
    'General Puerta Aqua':        'MUE008',
    'Puerta corrediza':           'MUE009',
    'Puerta Escalera Aqua':       'MUE010',
    'VESTIDOR DE RECAMARA PRINCIPAL': 'MUE011',
}

# Mapa mueble → proyecto
MUE_TO_PROY = {
    'MUE006': 'PROY002','MUE007': 'PROY002','MUE008': 'PROY002',
    'MUE009': 'PROY002','MUE010': 'PROY002','MUE011': 'PROY002',
}

# Salidas del Reporte Diario
SALIDAS_DIARIO = [
    ('CLOSET DE BLANCOS',   'Resistol 850',                                  'Federico Munoz',  1),
    ('Credeza',             'Clavillo C-23 - 5/8',                           'Martin Alvarez', 10),
    ('Credeza',             'Pivot Door Slide 8080PEZ 20',                   'Martin Alvarez',  2),
    ('General Puerta Aqua', 'Pija rapida 1 1/2 "',                           'Cruz',          100),
    ('General Puerta Aqua', 'Pija rapida 1 1/4 "',                           'Cruz',          100),
    ('General Puerta Aqua', 'Pija rapida 2 1/2 "',                           'Cruz',          100),
    ('General Puerta Aqua', 'Pija rapida 2"',                                'Cruz',          100),
    ('General Puerta Aqua', 'Pija rapida calibre 10 3"',                     'Cruz',          100),
    ('Puerta corrediza',    'Sistema colgante para puerta corrediza DN80CF', 'Pedro Briseno',   1),
    ('Puerta Escalera Aqua','Cinta Masking tape Tuck 3/4',                   'Carlos Ibarra',   1),
    ('Puerta Escalera Aqua','Cinta Masking tape tuck 3/4 Azul',              'Carlos Ibarra',   1),
    ('Puerta Escalera Aqua','Lija de agua grano 220',                        'Carlos Ibarra',   2),
    ('Puerta Escalera Aqua','Lija de agua grano 320',                        'Carlos Ibarra',   1),
    ('Puerta Escalera Aqua','Lija Orbital G120',                             'Carlos Ibarra',   4),
    ('Puerta Escalera Aqua','Lija Orbital G220',                             'Carlos Ibarra',   4),
    ('Puerta Escalera Aqua','Lija para madera grano 80',                     'Carlos Ibarra',   1),
    ('VESTIDOR DE RECAMARA PRINCIPAL','Lija Orbital G120',                   'Jesus Valdez',    1),
]

mov_hdrs = ['Fecha','Codigo_mat','Tipo','Cantidad','Unidad','ID_Proyecto','Mueble','ID_Trab','ID_Ubic_Origen','ID_Ubic_Destino']
fp3 = os.path.join(REG, 'RegistroMovMaterial.xlsx')
wb3 = load_workbook(fp3)
ws3 = wb3['MOV_ALMACEN']
clear_data(ws3)

count_mov = 0

# Entradas de inventario inicial
for cod, qty in STOCK_INICIAL.items():
    if qty > 0:
        row = {
            'Fecha':          '2026-01-01',
            'Codigo_mat':     cod,
            'Tipo':           'Entrada',
            'Cantidad':       qty,
            'Unidad':         UNIDADES.get(cod,'pza'),
            'ID_Proyecto':    '',
            'Mueble':         'Inventario inicial',
            'ID_Trab':        '',
            'ID_Ubic_Origen': '',
            'ID_Ubic_Destino':'',
        }
        append(ws3, row, mov_hdrs)
        count_mov += 1

# Salidas del Reporte Diario Almacen
for mue_name, mat_desc, trab, qty in SALIDAS_DIARIO:
    cod = DESC_TO_CODE.get(mat_desc)
    if not cod:
        print(f"  WARNING Sin código para: {mat_desc}")
        continue
    mue_id  = MUE_NAME_MAP.get(mue_name, '')
    proy_id = MUE_TO_PROY.get(mue_id, '')
    row = {
        'Fecha':          '2026-03-14',
        'Codigo_mat':     cod,
        'Tipo':           'Salida',
        'Cantidad':       qty,
        'Unidad':         UNIDADES.get(cod,'pza'),
        'ID_Proyecto':    proy_id,
        'Mueble':         mue_name,
        'ID_Trab':        trab,
        'ID_Ubic_Origen': '',
        'ID_Ubic_Destino':'',
    }
    append(ws3, row, mov_hdrs)
    count_mov += 1

wb3.save(fp3)
print(f"✓ MOV_ALMACEN      → {count_mov} movimientos ({len(STOCK_INICIAL)} entradas iniciales + {len(SALIDAS_DIARIO)} salidas)")
print("\n✅ Migración completada correctamente.")
